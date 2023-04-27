def create_yearly_zonal_workbook(in_variable, in_start_year, in_end_year):
    from openpyxl import Workbook, load_workbook
    import datetime
    rain_final = Workbook()
    year = in_start_year
    while year < in_end_year + 1:
        # defining dates
        start_day = datetime.date(year, 1, 1)
        end_day = datetime.date(year, 12, 31)
        timeDelta = datetime.timedelta(days=1)

        # create worksheets
        rain_final.create_sheet(f'{year}{in_variable}')
        rain_final_ws = rain_final[f'{year}{in_variable}']
        # column dimensions
        i = 0
        while i < 7:
            rain_final_ws.column_dimensions[chr(65 + i)].width = 14
            i += 1

        zone = 0
        while zone < 5:
            rain_final_ws[f'{chr(67 + zone)}1'] = f'Zone_{zone + 1}'
            zone += 1

        i = 0
        current_day = start_day
        while current_day < end_day + timeDelta:
            rain_final_ws[f'A{2 + i}'] = i + 1
            rain_final_ws[f'B{2 + i}'] = current_day
            i += 1
            current_day += timeDelta
        year += 1
    return rain_final


from openpyxl import Workbook, load_workbook
import pandas as pd
import datetime
import os

variable = input('Input variable(precip/tmin/tmax): ')
start_year = int(input("Initial year: "))
end_year = int(input('Final year: '))
wb_out = load_workbook(r"C:\Users\Lenovo\Desktop\Main Project\Z1. Data\test.xlsx")

current_year = start_year
while current_year < end_year + 1:
    variableLoc = r'C:\Users\Lenovo\Desktop\Main Project\Z1. Data\{0}\{1}zones'.format(variable, current_year)
    ws_out = wb_out[f'{current_year}{variable}']
    zone = 0
    while zone < 5:
        excel_file = f'zone_{zone+1}.xlsx'
        excel_file_path = os.path.join(variableLoc, excel_file)

        pd_in = pd.read_excel(excel_file_path)

        zonal_avg = list(pd_in.iloc[:, 2:].mean(axis=1))

        i = 0
        while i < len(zonal_avg):
            ws_out[f'{chr(67+zone)}{2+i}'] = zonal_avg[i]
            i += 1

        print (f"Zone {zone+1} of year {current_year} tabulated")
        zone += 1

    print(f'{current_year} data tabulated')
    current_year += 1

wb_out.save(r"C:\Users\Lenovo\Desktop\Main Project\Z1. Data\test2.xlsx")
print('Task completed successfully')
