variable = input('Input variable(precip/tmin/tmax): ')
start_year = int(input("Initial year: "))
end_year = int(input('Final year: '))

working_directory = input("Input main working directory: ")
total_zones = int(input("Total number of zones: "))


def create_yearly_zonal_workbook():
    from openpyxl import Workbook
    import datetime
    rain_final = Workbook()
    current_year = start_year
    while current_year < end_year + 1:
        # defining dates
        start_day = datetime.date(current_year, 1, 1)
        end_day = datetime.date(current_year, 12, 31)
        timeDelta = datetime.timedelta(days=1)

        # create worksheets
        rain_final.create_sheet(f'{current_year}{variable}')
        rain_final_ws = rain_final[f'{current_year}{variable}']
        # column dimensions
        i = 0
        while i < total_zones + 2:
            rain_final_ws.column_dimensions[chr(65 + i)].width = 14
            i += 1

        zone = 0
        while zone < total_zones:
            rain_final_ws[f'{chr(67 + zone)}1'] = f'Zone_{zone + 1}'
            zone += 1

        i = 0
        current_day = start_day
        while current_day < end_day + timeDelta:
            rain_final_ws[f'A{2 + i}'] = i + 1
            rain_final_ws[f'B{2 + i}'] = current_day
            i += 1
            current_day += timeDelta
        current_year += 1
    print(f"Blank table format of {variable} data created.")
    return rain_final


def create_zonal_rainfall_table():
    from openpyxl import Workbook, load_workbook
    import pandas as pd
    import datetime
    import os

    wb_out = create_yearly_zonal_workbook()

    current_year = start_year
    while current_year < end_year + 1:
        variableLoc = os.path.join(working_directory, r'{0}\{1}zones'.format(variable, current_year))
        ws_out = wb_out[f'{current_year}{variable}']
        zone = 0
        while zone < total_zones:
            excel_file = f'zone_{zone + 1}.xlsx'
            excel_file_path = os.path.join(variableLoc, excel_file)

            pd_in = pd.read_excel(excel_file_path)

            zonal_avg = list(pd_in.iloc[:, 2:].mean(axis=1))

            i = 0
            while i < len(zonal_avg):
                ws_out[f'{chr(67 + zone)}{2 + i}'] = zonal_avg[i]
                i += 1

            print(f"Zone {zone + 1} of year {current_year} tabulated")
            zone += 1

        print(f'{current_year} data tabulated')
        current_year += 1

    wb_out.save(os.path.join(working_directory, f'{variable}_data.xlsx'))
    print(f'Annual {variable} data tabulated zone-wise')


create_zonal_rainfall_table()
